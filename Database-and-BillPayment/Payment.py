import openpyxl
import os
import datetime
class Payment:
    '''The Payment class includes the following methods:
search_payment(PaymentID): Searches for a payment record by its ID.
search_payment_by_customer_id(CustomerID): Searches for payment records by a customer ID.
create_payment(): Creates a new payment record and appends it to the Payments.csv file.
The user inputs the CustomerID, BillingID, Payment Date, and Payment Amount.
The method checks if the BillingID and CustomerID match and if the Payment Amount matches the TotalBill in the Billing.csv file.'''

    def __init__(self, payment_file, billing_file):
        self.payment_file = payment_file
        self.billing_file = billing_file

        if not os.path.exists(self.payment_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["PaymentID", "BillingID", "CustomerID", "PaymentDate", "Month", "Year", "PaymentAmount"])
            wb.save(self.payment_file)

    def search_payment(self, payment_id):
        wb = openpyxl.load_workbook(self.payment_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == payment_id:
                data.append(row)
        return data

    def search_payment_by_customer_id(self, customer_id):
        wb = openpyxl.load_workbook(self.payment_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == customer_id:
                data.append(row)
        return data

    def create_payment(self, customer_id, billing_id, payment_date, month, year, payment_amount):
        wb_billing = openpyxl.load_workbook(self.billing_file)
        ws_billing = wb_billing.active

        valid_billing = False
        for row in ws_billing.iter_rows(min_row=2, values_only=True):
            if row[0] == billing_id and row[1] == customer_id:
                total_bill = row[8]
                valid_billing = True
                break

        if not valid_billing:
            print("Billing ID and Customer ID do not match. Cannot create payment.")
            return

        if payment_amount != total_bill:
            print("Payment amount does not match the total bill. Cannot create payment.")
            return

        wb_payment = openpyxl.load_workbook(self.payment_file)
        ws_payment = wb_payment.active
        payment_id = ws_payment.max_row - 1
        new_entry = [payment_id, billing_id, customer_id, payment_date, month, year, payment_amount]
        ws_payment.append(new_entry)
        wb_payment.save(self.payment_file)