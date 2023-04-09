import openpyxl
import os
import Bill_Calculate

class Billing:
    '''The Billing class now includes the following methods:
search_billing(BillingID): Searches for a billing record by its ID.
search_billing_by_customer_id(CustomerID): Searches for billing records by a customer ID.
auto_calculate_billing(): Automatically calculates billing amounts and appends the data to the Billing.csv file.
auto_check_status_and_update(): Automatically checks the billing status and updates the late fee and status in the Billing.csv file.'''

    def __init__(self, meter_reading_file):
        self.meter_reading_file = meter_reading_file
        if not os.path.exists(self.meter_reading_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["MeterReadingID", "CustomerID", "Time", "Date", "Month", "Year", "ReadingAmount"])
            wb.save(self.meter_reading_file)

    def create_meter_reading(self, customer_id):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        meter_reading_id = ws.max_row
        new_entry = [meter_reading_id, customer_id, "", "", "", "", ""]
        ws.append(new_entry)
        wb.save(self.meter_reading_file)
        return meter_reading_id

    def input_data_reading(self, meter_reading_id, hour, date, month, year, reading_amount):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == meter_reading_id:
                found = True
                ws.cell(row=row[0], column=3, value=hour)
                ws.cell(row=row[0], column=4, value=date)
                ws.cell(row=row[0], column=5, value=month)
                ws.cell(row=row[0], column=6, value=year)
                ws.cell(row=row[0], column=7, value=reading_amount)
                break
        if not found:
            print("MeterReadingID not found")
        else:
            wb.save(self.meter_reading_file)

    def search_meter_reading(self, meter_reading_id):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == meter_reading_id:
                data.append(row)
        return data

    def search_meter_reading_by_customer_id(self, customer_id):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == customer_id:
                data.append(row)
        return data

    def update_meter_reading(self, meter_reading_id, operation):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        if operation.lower() == "delete":
            rows_to_delete = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == meter_reading_id:
                    rows_to_delete.append(row)
            for row in rows_to_delete:
                ws.delete_rows(row[0])
            wb.save(self.meter_reading_file)
        elif operation.lower() == "modify":
            print("Enter new data for the meter reading")
            hour = input("Hour: ")
            date = input("Date: ")
            month = input("Month: ")
            year = input("Year: ")
            reading_amount = float(input("Reading Amount: "))
            self.input_data_reading(meter_reading_id, hour, date, month, year, reading_amount)
        else:
            print("Invalid operation. Please choose 'delete' or 'modify'.")

def main():
    meter_reading_file = "MeterReading.xlsx"
    meter_reading = MeterReading(meter_reading_file)

    while True:
        print("\n--- Meter Reading Management ---")
        print("1. Create Meter Reading")
        print("2. Input Data Reading")
        print("3. Search Meter Reading by ID")
        print("4. Search Meter Reading by Customer ID")
        print("5. Update Meter Reading")
        print("6. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            customer_id = int(input("Enter Customer ID: "))
            meter_reading.create_meter_reading(customer_id)
            print("Meter Reading created successfully.")
        elif choice == "2":
            meter_reading_id = int(input("Enter Meter Reading ID: "))
            hour = input("Enter Hour: ")
            date = input("Enter Date: ")
            month = input("Enter Month: ")
            year = input("Enter Year: ")
            reading_amount = float(input("Enter Reading Amount: "))
            meter_reading.input_data_reading(meter_reading_id, hour, date, month, year, reading_amount)
            print("Data reading added successfully.")
        elif choice == "3":
            meter_reading_id = int(input("Enter Meter Reading ID: "))
            results = meter_reading.search_meter_reading(meter_reading_id)
            if results:
                print("\nResults:")
                print("MeterReadingID, CustomerID, Time, Date, Month, Year, ReadingAmount")
                for result in results:
                    print(result)
            else:
                print("No matching Meter Reading ID found.")
        elif choice == "4":
            customer_id = int(input("Enter Customer ID: "))
            results = meter_reading.search_meter_reading_by_customer_id(customer_id)
            if results:
                print("\nResults:")
                print("MeterReadingID, CustomerID, Time, Date, Month, Year, ReadingAmount")
                for result in results:
                    print(result)
            else:
                print("No matching Customer ID found.")
        elif choice == "5":
            meter_reading_id = int(input("Enter Meter Reading ID: "))
            operation = input("Choose operation: delete or modify: ")
            meter_reading.update_meter_reading(meter_reading_id, operation)
            print("Meter Reading updated successfully.")
        elif choice == "6":
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()