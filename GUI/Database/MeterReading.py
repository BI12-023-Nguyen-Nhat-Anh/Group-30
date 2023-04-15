import openpyxl
import os


class MeterReading:
    def __init__(self, meter_reading_file):
        self.meter_reading_file = meter_reading_file
        if not os.path.exists(self.meter_reading_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["MeterReadingID", "CustomerID", "Time", "Date", "Month", "Year", "ReadingAmount"])
            wb.save(self.meter_reading_file)

    def create_meter_reading(self, customer_id):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        meter_reading_id = ws.max_row
        new_entry = [meter_reading_id, customer_id, "", "", "", "", ""]
        ws.append(new_entry)
        wb.save(self.meter_reading_file)
        return meter_reading_id

    def input_data_reading(self, meter_reading_id, hour, date, month, year, reading_amount):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == meter_reading_id:
                found = True
                ws.cell(row=row[0], column=3, value=hour)
                ws.cell(row=row[0], column=4, value=date)
                ws.cell(row=row[0], column=5, value=month)
                ws.cell(row=row[0], column=6, value=year)
                ws.cell(row=row[0], column=7, value=reading_amount)
                break
        if not found:
            print("MeterReadingID not found")
        else:
            wb.save(self.meter_reading_file)

    def search_meter_reading(self, meter_reading_id):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == meter_reading_id:
                data.append(row)
        return data

    def search_meter_reading_by_customer_id(self, customer_id):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == customer_id:
                data.append(row)
        return data

    def update_meter_reading(self, meter_reading_id, operation):
        wb = openpyxl.load_workbook(self.meter_reading_file)
        ws = wb.active
        if operation.lower() == "delete":
            rows_to_delete = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == meter_reading_id:
                    rows_to_delete.append(row)
            for row in rows_to_delete:
                ws.delete_rows(row[0])
            wb.save(self.meter_reading_file)
        elif operation.lower() == "modify":
            print("Enter new data for the meter reading")
            hour = input("Hour: ")
            date = input("Date: ")
            month = input("Month: ")
            year = input("Year: ")
            reading_amount = float(input("Reading Amount: "))
            self.input_data_reading(meter_reading_id, hour, date, month, year, reading_amount)
        else:
            print("Invalid operation. Please choose 'delete' or 'modify'.")