import openpyxl
import os
import datetime

class Billing:
    '''The Billing class now includes the following methods:
search_billing(BillingID): Searches for a billing record by its ID.
search_billing_by_customer_id(CustomerID): Searches for billing records by a customer ID.
auto_calculate_billing(): Automatically calculates billing amounts and appends the data to the Billing.xlsx file.
auto_check_status_and_update(): Automatically checks the billing status and updates the late fee and status in the Billing.xlsx file.'''

    def __init__(self, billing_file, consumption_file, customer_file):
        self.billing_file = billing_file
        self.consumption_file = consumption_file
        self.customer_file = customer_file

        if not os.path.exists(self.billing_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["BillingID", "CustomerID", "ConsumptionID", "BillingDeadline", "Month", "Year", "BillingAmount",
                       "LateFee", "TotalBill", "Status"])
            wb.save(self.billing_file)

    def search_billing(self, billing_id):
        wb = openpyxl.load_workbook(self.billing_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == billing_id:
                data.append(row)
        return data

    def search_billing_by_customer_id(self, customer_id):
        wb = openpyxl.load_workbook(self.billing_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == customer_id:
                data.append(row)
        return data

    def auto_calculate_billing(self):
        customer_types = {}
        wb_customer = openpyxl.load_workbook(self.customer_file)
        ws_customer = wb_customer.active
        for row in ws_customer.iter_rows(min_row=2, values_only=True):
            customer_id = row[0]
            customer_type = row[7]
            customer_types[customer_id] = customer_type

        wb_consumption = openpyxl.load_workbook(self.consumption_file)
        ws_consumption = wb_consumption.active
        for row in ws_consumption.iter_rows(min_row=2, values_only=True):
            consumption_id = row[0]
            customer_id = row[1]
            month = row[2]
            year = row[3]
            consumption_amount = row[4]

            customer_type = customer_types.get(customer_id)
            if customer_type is None:
                print(f"Customer type not found for CustomerID {customer_id}. Skipping this record.")
                continue

            billing_amount = 0
            if customer_type == "Household":
                billing_amount = Household(consumption_amount)
            elif customer_type == "Manufacturing_industries":
                billing_amount = Manufacturing_industries(consumption_amount)
            elif customer_type == "Administrative_offices":
                billing_amount = Administrative_offices(consumption_amount)
            elif customer_type == "Business":
                billing_amount = Business(consumption_amount)

            deadline_month = month + 1
            deadline_year = year
            if deadline_month > 12:
                deadline_month = 1
                deadline_year += 1

            billing_deadline = f"{deadline_year}-{deadline_month:02d}-05"
            late_fee = 0
            total_bill = billing_amount + late_fee
            status = "Pending"

            self._append_billing_data(consumption_id, customer_id, billing_deadline, month, year, billing_amount,
                                      late_fee, total_bill, status)

    def _append_billing_data(self, consumption_id, customer_id, billing_deadline, month, year, billing_amount, late_fee,
                             total_bill, status):
        wb = openpyxl.load_workbook(self.billing_file)
        ws = wb.active
        billing_id = ws.max_row - 1
        new_entry = [billing_id, customer_id, consumption_id, billing_deadline, month, year, billing_amount, late_fee,
                     total_bill, status]
        ws.append(new_entry)
        wb.save(self.billing_file)

    def auto_check_status_and_update(self):
        today = datetime.date.today()
        wb_billing = openpyxl.load_workbook(self.billing_file)
        ws_billing = wb_billing.active
        updated = False

        for row in ws_billing.iter_rows(min_row=2):
            status = row[9].value
            if status == "Paid":
                continue

            billing_deadline = datetime.datetime.strptime(row[3].value, "%Y-%m-%d").date()
            if today > billing_deadline:
                billing_amount = row[6].value
                late_fee = billing_amount * 0.1
                total_bill = billing_amount + late_fee
                row[7].value = late_fee
                row[8].value = total_bill
                row[9].value = "Overdue"
                updated = True
        if updated:
            wb_billing.save(self.billing_file)
