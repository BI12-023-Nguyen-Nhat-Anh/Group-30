import tkinter as tk
import openpyxl
from tkinter import ttk

# Initialize root window
root = tk.Tk()
root.title('Billing Information')
root.pack_propagate(False)
root.geometry("1280x720")
root.resizable(0, 0)

# Configure theme
style = ttk.Style(root)
theme_path = "assets/forest-dark.tcl"
root.tk.call("source", theme_path)
style.theme_use("forest-dark")

"""
=================TREEVIEW======================================
===============================================================
"""
# Create paned window
paned = ttk.PanedWindow(root)
paned.grid(row=0, column=0, columnspan=10, padx=(10, 20),
           pady=(10, 10), sticky="nsew")
paned.pack_propagate(False)

# Pane #1
pane_1 = ttk.Frame(paned)
paned.add(pane_1, weight=1)

# tree frame
treeFrame = ttk.Frame(pane_1)
treeFrame.grid(row=0, column=0, padx=5, pady=5)

# scroll ball
treeScrolly = ttk.Scrollbar(treeFrame)
treeScrolly.pack(side="right", fill="y")

# contents
cols = ("BillingID", "CustomerID", "ConsumptionID",
        "BillingDeadline", "BillingAmount", "LateFee", "TotalBill", "Status")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScrolly.set, columns=cols, height=13)

for col in cols:
    treeview.heading(col, text=col)
    treeview.column(col, width=150, anchor="center")
treeview.pack(fill="both", expand=True)
treeScrolly.config(command=treeview.yview)

path = "data/Billing.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

# Load data into treeview
def load_data():
    global list_values
    list_values = list(sheet.values)[1:]
    for value_tuple in list_values:
        treeview.insert("", tk.END, values=value_tuple[:8])

"""
=================Search and filter======================================
===============================================================
"""
# Search and filter
search_frame = ttk.LabelFrame(
    root, text="Search and filter", padding=(20, 10))
search_frame.grid(row=1, column=0, padx=(15, 10),
                  pady=(10, 10), columnspan=2, sticky="nsew")
search_frame.columnconfigure(index=0, weight=1)
search_frame.pack_propagate(False)

# CustomerID entry
customer_id_entry = ttk.Entry(search_frame)
customer_id_entry.insert(0, "CustomerID")
# double left-click on entry and the default text disappears, right-click to get the default text
customer_id_entry.bind("<Double-Button-1>", lambda e: customer_id_entry.delete(0, "end"))
customer_id_entry.bind("<Button-3>", lambda e: customer_id_entry.insert(
    0, "Customer Code") if not customer_id_entry.get() else None)
customer_id_entry.grid(row=1, column=0, columnspan=2, pady=10, sticky="ew")

# BillingID entry
billing_id_entry = ttk.Entry(search_frame)
billing_id_entry.insert(0, "BillingID")
billing_id_entry.bind("<Double-Button-1>", lambda e: billing_id_entry.delete(0, "end"))
billing_id_entry.bind("<Button-3>", lambda e: billing_id_entry.insert(
    0, "Name") if not billing_id_entry.get() else None)
billing_id_entry.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")


# Status
status_list = ["Pending", "Paid", "Overdue"]
status_combo = ttk.Combobox(search_frame, state="readonly", values=status_list)
status_combo.grid(row=4, column=0, pady=10, sticky="ew")
# right-click to clear the option
status_combo.bind("<Button-3>", lambda event: clear_selection_type(event))

def clear_selection_type(event):
    type_combo.selection_clear()
    type_combo.set("Status")
# Reset button
reset_button = ttk.Button(search_frame, text="Reset", command=lambda: reset())
reset_button.grid(row=5, column=1, pady=15, sticky="nsew")
reset_button.bind("<Button-1>", lambda event: reset(event))

# Search button
search_button = ttk.Button(search_frame, text="Search", style="Accent.TButton", command=lambda: search())
search_button.grid(row=5, column=0, pady=15, padx=(0, 10), sticky="nsew")
search_button.bind("<Button-1>", lambda event: search(event))

def search(event):
    # Get the values from the search inputs
    customer_id_value = customer_id_entry.get()
    billing_id_value = billing_id_entry.get()
    status_value = status_combo.get()

    # Clear the treeview
    treeview.delete(*treeview.get_children())

    # Filter the data based on the search inputs
    filtered_data = []
    for value_tuple in list_values:
        if ((customer_id_value == "CustomerID" or customer_id_value in value_tuple[1]) and
            (billing_id_value == "BillingID" or billing_id_value in value_tuple[0])):
            if status_value == "Status":
                filtered_data.append(value_tuple[:8])
            elif status_value in value_tuple[7]:
                filtered_data.append(value_tuple[:8])

    # Update the treeview with the filtered data
    for value_tuple in filtered_data:
        treeview.insert("", tk.END, values=value_tuple)

def reset(event):
    customer_id_entry.delete(0, "end")
    customer_id_entry.insert(0, "CustomerID")
    billing_id_entry.delete(0, "end")
    billing_id_entry.insert(0, "BillingID")
    status_combo.set("Status")
    treeview.delete(*treeview.get_children())

    # ensure a row of headers is not added in the treeview
    for value_tuple in list_values:
        treeview.insert("", tk.END, values=value_tuple[:8])


"""
=================Notebook======================================
===============================================================
"""

paned2 = ttk.PanedWindow(root)
paned2.grid(row=1, column=2, padx=(10, 0),
            pady=(18, 0), columnspan=7, sticky="nsew")
# Pane #2
pane_2 = ttk.Frame(paned)
paned.add(pane_2, weight=2)
paned2.pack_propagate(False)

# Notebook
notebook = ttk.Notebook(pane_2, height=280)
notebook.pack(fill=tk.BOTH, expand=True)

# Tab #1
tab_1 = ttk.Frame(notebook)
notebook.add(tab_1, text="Details")

def on_treeview_select(event):
    selection = event.widget.selection()
    if selection:
        selected_item = treeview.focus()
        chosen_id = treeview.item(selected_item)['values'][0]
        path = "data/Billing.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        cols = ("BillingID", "CustomerID", "ConsumptionID", "BillingDeadline", "BillingAmount", "LateFee", "TotalBill", "Status")
        row = None
        for r in sheet.iter_rows(min_row=2):
            if r[0].value == chosen_id:
                row = r
                break

        if row:
            for widget in tab_1.winfo_children():
                widget.destroy()

            for j, header in enumerate(cols):
                header = cols[j]
                ttk.Label(tab_1, text=header + ":").grid(
                    column=0, row=j, sticky="w", pady=5, padx=10)
                ttk.Label(tab_1, text=row[j].value).grid(
                    column=1, row=j, sticky="w", pady=5, padx=10)

treeview.bind('<<TreeviewSelect>>', on_treeview_select)

# center window
root.update()
root.minsize(root.winfo_width(), root.winfo_height())
x_cordinate = int((root.winfo_screenwidth()/2) - (root.winfo_width()/2))
y_cordinate = int((root.winfo_screenheight()/2) - (root.winfo_height()/2))
root.geometry("+{}+{}".format(x_cordinate, y_cordinate))

load_data()
root.mainloop()