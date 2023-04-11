import tkinter as tk
import openpyxl
from tkinter import ttk
import os

root = tk.Tk()
root.title('Customer list')
root.option_add("*tearOff", False)
root.pack_propagate(False)
root.geometry("1280x720")
root.resizable(0, 0)
# Create a style
style = ttk.Style(root)
# Import the tcl file
theme_path = os.path.join(os.path.dirname(
    __file__), '..', 'assets', 'forest-dark.tcl')

# Load the theme file
root.tk.call("source", theme_path)

# Set the theme with the theme_use method
style.theme_use("forest-dark")


"""
=================TREEVIEW======================================
===============================================================
"""
# Panedwindow
paned = ttk.PanedWindow(root)
paned.grid(row=0, column=0, columnspan=10, padx=(10, 20),
           pady=(10, 10), sticky="nsew")
paned.pack_propagate(False)

# Pane #1
pane_1 = ttk.Frame(paned)
paned.add(pane_1, weight=1)


# tree frame
treeFrame = ttk.Frame(pane_1)
treeFrame.grid(row=0, column=0, padx=5, pady=5)


# scroll ball
treeScrolly = ttk.Scrollbar(treeFrame)
treeScrolly.pack(side="right", fill="y")

# contents
cols = ("Customer Code", "Name", "Electricity usage address", "Phone Number",
        "Identity number", "Tax Code", "Type", "Status")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScrolly.set, columns=cols, height=13)

treeview.column("Customer Code", width=120, anchor="center")
treeview.column("Name", width=140)
treeview.column("Electricity usage address", width=330)
treeview.column("Phone Number", width=120, anchor="center")
treeview.column("Identity number", width=120, anchor="center")
treeview.column("Tax Code", width=120, anchor="center")
treeview.column("Type", width=160, anchor="center")
treeview.column("Status", width=90, anchor="center")
treeview.bind('<Button-1>', lambda event: on_treeview_select(event))

treeview.pack(fill="both", expand=True)
treeScrolly.config(command=treeview.yview)

path = os.path.join(os.path.dirname(__file__), '..',
                    'data', 'data_customer.xlsx')
workbook = openpyxl.load_workbook(path)
sheet = workbook["filtered_data"]


def load_data():
    global list_values
    list_values = list(sheet.values)
    for col_name in list_values[0]:
        if col_name in cols:
            treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert("", tk.END, values=value_tuple)


"""
=================Search and filter======================================
===============================================================
"""
# search & filter
search_frame = ttk.LabelFrame(
    root, text="Search and filter", padding=(20, 10))
search_frame.grid(row=1, column=0, padx=(15, 10),
                  pady=(10, 10), columnspan=2, sticky="nsew")
search_frame.columnconfigure(index=0, weight=1)
search_frame.pack_propagate(False)
# ID
id_entry = ttk.Entry(search_frame)
id_entry.insert(0, "Customer Code")
# double left-click on entry and the default text disappears, right-click to get the default text
id_entry.bind("<Double-Button-1>", lambda e: id_entry.delete(0, "end"))
id_entry.bind("<Button-3>", lambda e: id_entry.insert(
    0, "Customer Code") if not id_entry.get() else None)
id_entry.grid(row=1, column=0, columnspan=2, pady=10, sticky="ew")


# Name
name_entry = ttk.Entry(search_frame)
name_entry.insert(0, "Name")
name_entry.bind("<Double-Button-1>", lambda e: name_entry.delete(0, "end"))
name_entry.bind("<Button-3>", lambda e: name_entry.insert(
    0, "Name") if not name_entry.get() else None)
name_entry.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")

# Type
type_list = ["Manufacturing industry",
             "Administrative office", "Business", "Household"]
type_combo = ttk.Combobox(search_frame, state="readonly", values=type_list)
type_combo.set("Type")
# right-click to clear the option
type_combo.bind("<Button-3>", lambda event: clear_selection_type(event))


def clear_selection_type(event):
    type_combo.selection_clear()
    type_combo.set("Type")


type_combo.grid(row=3, column=0, columnspan=2, pady=10, sticky="ew")

# Status
status_list = ["Active", "Inactive"]
status_combo = ttk.Combobox(search_frame, state="readonly", values=status_list)
status_combo.set("Status")
status_combo.bind("<Button-3>", lambda event: clear_selection_status(event))


def clear_selection_status(event):
    status_combo.selection_clear()
    status_combo.set("Status")


status_combo.grid(row=4, column=0, columnspan=2, pady=10, sticky="ew")

# Reset button
reset_button = ttk.Button(search_frame, text="Reset")
reset_button.grid(row=5, column=1, pady=15, sticky="nsew")
reset_button.bind("<Button-1>", lambda event: reset(event))

# Search button
search_button = ttk.Button(search_frame, text="Search", style="Accent.TButton")
search_button.grid(row=5, column=0, pady=15, padx=(0, 10), sticky="nsew")
search_button.bind("<Button-1>", lambda event: search(event))


def search(event):
    # Get the values from the search inputs
    id_value = id_entry.get()
    name_value = name_entry.get().lower()
    type_value = type_combo.get()
    status_value = status_combo.get()

    # Clear the treeview
    treeview.delete(*treeview.get_children())

    # Filter the data based on the search inputs
    filtered_data = []
    for value_tuple in list_values[1:]:
        if (id_value == "Customer Code" or id_value in value_tuple[0]) or (name_value == "Name" or name_value in value_tuple[1].lower() and (all(word in value_tuple[1].lower() for word in name_value))):
            if type_value == "Type" and status_value == "Status":
                filtered_data.append(value_tuple)
            elif type_value != "Type" and status_value == "Status":
                if type_value in value_tuple[6]:
                    filtered_data.append(value_tuple)
            elif type_value == "Type" and status_value != "Status":
                if status_value in value_tuple[7]:
                    filtered_data.append(value_tuple)
            elif type_value != "Type" and status_value != "Status":
                if type_value in value_tuple[6] and status_value in value_tuple[7]:
                    filtered_data.append(value_tuple)

    # Update the treeview with the filtered data
    for value_tuple in filtered_data:
        treeview.insert("", tk.END, values=value_tuple)


def reset(event):
    id_entry.delete(0, "end")
    id_entry.insert(0, "Customer Code")
    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    type_combo.set("Type")
    status_combo.set("Status")
    treeview.delete(*treeview.get_children())
    # ensure a row of headers is not added in the treeview
    for value_tuple in list_values[1:]:
        treeview.insert("", tk.END, values=value_tuple)


"""
=================Notebook======================================
===============================================================
"""

paned2 = ttk.PanedWindow(root)
paned2.grid(row=1, column=2, padx=(10, 0),
            pady=(18, 0), columnspan=7, sticky="nsew")
# Pane #2
pane_2 = ttk.Frame(paned2)
paned2.add(pane_2, weight=2)
paned2.pack_propagate(False)

# Notebook
notebook = ttk.Notebook(pane_2, height=280)
notebook.pack(fill=tk.BOTH, expand=True)


# Tab #1
tab_1 = ttk.Frame(notebook)
notebook.add(tab_1, text="Details")


def on_treeview_select(event):
    selection = event.widget.selection()
    if selection:
        selected_item = treeview.focus()
        chosen_id = treeview.item(selected_item)['values'][0]
        path = os.path.join(os.path.dirname(__file__), '..',
                            'data', 'data_customer.xlsx')
        workbook = openpyxl.load_workbook(path)
        sheet2 = workbook["data_customer"]
        cols2 = ("Customer Code", "Name", "Electricity usage address", "Residential address", "Phone Number", "Email",
                 "Identity number", "Tax Code", "Type", "Status")
        row = None
        for r in sheet2.iter_rows(min_row=2):
            if r[0].value == chosen_id:
                row = r
                break

        if row:
            # Clear the current widgets in tab_1
            for widget in tab_1.winfo_children():
                widget.destroy()

            # Create labels for each header/value pair
            for j, header in enumerate(cols2):
                header = cols2[j]
                ttk.Label(tab_1, text=header + ":").grid(
                    column=0, row=j, sticky="w", pady=5, padx=10)
                ttk.Label(tab_1, text=row[j].value).grid(
                    column=1, row=j, sticky="w", pady=5, padx=10)


# center window
root.update()
root.minsize(root.winfo_width(), root.winfo_height())
x_cordinate = int((root.winfo_screenwidth()/2) - (root.winfo_width()/2))
y_cordinate = int((root.winfo_screenheight()/2) - (root.winfo_height()/2))
root.geometry("+{}+{}".format(x_cordinate, y_cordinate))

load_data()
root.mainloop()
