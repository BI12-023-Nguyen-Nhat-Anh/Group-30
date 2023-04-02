from tkinter import *
from tkinter.font import Font
import openpyxl

root = Tk()
root.title('Dashboard')

# adjust size
root.geometry("1280x719")
# tells the root to not let the widgets inside it determine its size.
root.pack_propagate(False)
root.resizable(0, 0)  # makes the root window fixed in size.

# set the background image
img = PhotoImage(file="assets/Dashboard.png")

# create a canvas on top of the label to make it clickable
canvas = Canvas(root, width=img.width(), height=img.height())
canvas.place(x=0, y=0)

# display the image on the canvas
canvas.create_image(0, 0, image=img, anchor="nw")

# add a rectangle to the canvas to define the clickable region
rect1 = canvas.create_rectangle(16, 102, 158, 133, fill="", outline="")
canvas.tag_bind(rect1, "<Button-1>", lambda event: switch_window_customer())

rect2 = canvas.create_rectangle(16, 136, 158, 167, fill="", outline="")
canvas.tag_bind(rect2, "<Button-1>", lambda event: switch_window_bill())

# tạm để là click, sẽ thêm file sau


def switch_window_customer():
    print("Clicked Customer")


def switch_window_bill():
    print("Clicked bill")


# set the font
family_font = Font(family="Space Mono", size=30, weight="normal")

# access the excel
path = "D:\Group-30\data\Dashboard.xlsx"
workbook = openpyxl.load_workbook(path)
all_customer = workbook["Customer"]

# Get the number of customers in the worksheet
num_customer = all_customer.max_row - 1

# display number of customers
num_customer = Label(root, text=f"{num_customer}",
                     font=family_font, bg="#EAF5FE", fg="#333536")
num_customer.place(x=234.36, y=139.01, width=49, height=40)

cus_status_col = "E"

# display number of active customers
num_active_customer = 0
for row in all_customer.iter_rows(min_row=2, values_only=True):
    if row[all_customer[f"{cus_status_col}1"].column - 1].lower() == "active":
        num_active_customer += 1
num_active_customer = Label(root, text=f"{num_active_customer}",
                            font=family_font, bg="#F4F0FF", fg="#333536")
num_active_customer.place(x=594.36, y=139.01, width=49, height=40)

# display number of inactive customers
num_inactive_customer = 0
for row in all_customer.iter_rows(min_row=2, values_only=True):
    if row[all_customer[f"{cus_status_col}1"].column - 1].lower() == "inactive":
        num_inactive_customer += 1
num_inactive_customer = Label(root, text=f"{num_inactive_customer}",
                              font=family_font, bg="#FEF1EF", fg="#333536")
num_inactive_customer.place(x=954.36, y=143.01, width=49, height=40)

# access the Bill excel
all_bill = workbook["Bill"]

# Get the number of bills in the worksheet
num_bill = all_bill.max_row - 1

# display number of customers
num_bill = Label(root, text=f"{num_bill}",
                 font=family_font, bg="#FEF6E8", fg="#333536")
num_bill.place(x=234.36, y=447.01, width=64, height=49.67)

bill_status_col = "G"

# display number of paid bills
num_paid_bill = 0
for row in all_bill.iter_rows(min_row=2, values_only=True):
    if row[all_bill[f"{bill_status_col}1"].column - 1].lower() == "paid":
        num_paid_bill += 1
num_paid_bill = Label(root, text=f"{num_paid_bill}",
                      font=family_font, bg="#FEFBE5", fg="#333536")
num_paid_bill.place(x=594.36, y=447.01, width=64, height=49.67)

# display number of overdue bills
num_overdue_bill = 0
for row in all_bill.iter_rows(min_row=2, values_only=True):
    if row[all_bill[f"{bill_status_col}1"].column - 1].lower() == "overdue":
        num_overdue_bill += 1
num_overdue_bill = Label(root, text=f"{num_overdue_bill}",
                         font=family_font, bg="#E4FEF4", fg="#333536")
num_overdue_bill.place(x=954.36, y=447.01, width=64, height=49.67)

root.mainloop()