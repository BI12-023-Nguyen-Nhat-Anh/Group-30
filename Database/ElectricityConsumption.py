import openpyxl
import os
class ElectricityConsumption:
    '''This class includes the following methods:
search_consumption(ConsumptionID): Searches for a consumption record by its ID.
search_consumption_by_customer_id(CustomerID): Searches for consumption records by a customer ID.
auto_get_consumption(): Automatically calculates the electricity consumption amounts for each customer and appends the data to the ElectricityConsumption.xlsx file.'''

    def __init__(self, consumption_file, meter_reading_file):
        self.consumption_file = consumption_file
        self.meter_reading_file = meter_reading_file

        if not os.path.exists(self.consumption_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["ConsumptionID", "CustomerID", "Month", "Year", "ConsumptionAmount"])
            wb.save(self.consumption_file)

    def search_consumption(self, consumption_id):
        wb = openpyxl.load_workbook(self.consumption_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == consumption_id:
                data.append(row)
        return data

    def search_consumption_by_customer_id(self, customer_id):
        wb = openpyxl.load_workbook(self.consumption_file)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == customer_id:
                data.append(row)
        return data

    def auto_get_consumption(self):
        meter_readings = {}
        wb_mr = openpyxl.load_workbook(self.meter_reading_file)
        ws_mr = wb_mr.active

        for row in ws_mr.iter_rows(min_row=2, values_only=True):
            customer_id = row[1]
            if customer_id not in meter_readings:
                meter_readings[customer_id] = []
            meter_readings[customer_id].append(row)

        for customer_id, readings in meter_readings.items():
            readings.sort(key=lambda x: (x[5], x[4]))

            for i in range(len(readings) - 1):
                consumption_id = i + 1
                month = readings[i][4]
                year = readings[i][5]
                consumption_amount = readings[i + 1][6] - readings[i][6]

                self._append_consumption_data(consumption_id, customer_id, month, year, consumption_amount)

    def _append_consumption_data(self, consumption_id, customer_id, month, year, consumption_amount):
        wb = openpyxl.load_workbook(self.consumption_file)
        ws = wb.active
        new_entry = [consumption_id, customer_id, month, year, consumption_amount]
        ws.append(new_entry)
        wb.save(self.consumption_file)